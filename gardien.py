import os
import pandas as pd
import random
import numpy as np
import time
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from solve import solve_multi

MAX_DIST = 10

ascii_art = r"""

                                          
  ____    _    ____  ____ ___ _____ _   _ 
 / ___|  / \  |  _ \|  _ |_ _| ____| \ | |
| |  _  / _ \ | |_) | | | | ||  _| |  \| |
| |_| |/ ___ \|  _ <| |_| | || |___| |\  |
 \____/_/   \_|_| \_|____|___|_____|_| \_|
                                          

"""

couleur_gardien = '\033[35m'

def print_ascii():
    print(f"\033[1m\033[31m{ascii_art}\033[0m")

def main():
    dir = input(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Répertoire où se trouvent les fichiers Excel (.xlsx) : ")
    dir = dir.strip()
    dir = dir.replace("\\ ", " ")
    dir = os.path.abspath(dir)

    verbose = True

    while True:
        mode = input(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Remplacer les fichiers existants ou créer de Nouveaux fichiers ? (R/N) : ").strip().upper()
        if mode in ['R', 'N']:
            break
        print(f"\033[1m\033[31m[ERREUR]\033[0m Veuillez répondre par 'R' (remplacer) ou 'N' (nouveaux fichiers)")

    print()

    try:
        excel_files = [
            f for f in os.listdir(dir) 
            if f.endswith('.xlsx') # on prend seulement les fichiers Excel
            and not f.startswith('~')  # on exclue les fichiers temporaires (~$...)
            and not f.startswith('.')  # les fichiers cachés (.fichier)
            and not f.startswith('$')
        ]
        random.shuffle(excel_files)
        assert len(excel_files) > 0, f"Aucun fichier .xlsx trouvés dans {dir}. Chemin d'exécution: {os.getcwd()}"
    except FileNotFoundError:
        print(f"\033[1m\033[31m[ERREUR]\033[0m Répertoire non trouvé : \033[1m\033[33m{dir}\033[0m")
        return
    except PermissionError:
        print(f"\033[1m\033[31m[ERREUR]\033[0m Accès refusé au répertoire : \033[1m\033[33m{dir}\033[0m")
        return
    except Exception as e:
        print(f"\033[1m\033[31m[ERREUR]\033[0m Une erreur inattendue s'est produite : \033[1m\033[33m{str(e)}\033[0m")
        return

    if len(excel_files) == 1:
        print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m \033[1m{len(excel_files)}\033[0m fichier trouvé: \033[1m\033[33m{', '.join(excel_files)}\033[0m")
    else:
        print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m \033[1m{len(excel_files)}\033[0m fichiers trouvés: \033[1m\033[33m{', '.join(excel_files)}\033[0m")

    eqs_to_global = []
    global_to_eqs = []
    preferences_eqs = []
    reductions_eqs = []
    attributs_eqs: List[Dict[str, List[bool]]] = []
    implications_eqs = []
    Ns = []
    Ds = []
    mdc_eqs = []

    global_mdc = set()

    planning_initiaux = [] # stocke les valeurs initiales de chaque planning (si non vide)
    jours_a_modifier = [] # stocke les jours mis en gras (ie, à modifier)
    jours_fixes = [] # stocke les jours soulignés (ie, à fixer)
    doit_modifier = [] # pour chaque planning, dit si une optimisation doit être faite sur ce planning
    skip_optims = [] # idem (on skip l'optimisation si toutes les cases sont remplies et qu'aucune modif n'est demandée)

    for file in excel_files:
        df = pd.read_excel(os.path.join(dir, file)) # lecture première feuille
        df = df.drop(df.columns[:3], axis=1) # lecture des préférences : on enlève les 3 premières colonnes ("jour", "garde", "astreinte")
        df = df.fillna(0) # on remplace les cases vides par des 0

        mdc = df.columns.to_list()
        preferences = df.to_numpy().T
        N, D = preferences.shape

        wb = load_workbook(os.path.join(dir, file)) # on va parcourir la première feuille
        ws = wb.worksheets[0]

        gardes = [] # on collecte les valeurs dans la colonne "garde" (-1 si vide)
        astreintes = [] # idem avec les astreintes
        jours_gras = {'garde': [], 'astreinte': []} # stocker les jours en gros (à modifier)
        jours_soulignes = {'garde': [], 'astreinte': []} # stocker les jours soulignés (à fixer)
        
        # parcours de la colonne "jour"
        # si on détecte un jour souligné, on ajoute à jours_soulignes toutes les gardes et astreintes précédentes
        for row in range(2, ws.max_row + 1):
            if row == D+2:
                break

            if ws.cell(row=row, column=1).font.underline:
                jours_soulignes['garde'].extend(range(row - 1))
                jours_soulignes['garde'] = list(set(jours_soulignes['garde'])) # on évite ici les doublons en passant par set() puis list()
                jours_soulignes['astreinte'].extend(range(row - 1))
                jours_soulignes['astreinte'] = list(set(jours_soulignes['astreinte']))
                break
        
        # parcours des colonnes "garde" et "astreinte"
        for row in range(2, ws.max_row + 1):
            if row == D+2:
                break

            garde = ws.cell(row=row, column=2).value
            astreinte = ws.cell(row=row, column=3).value
            
            # check si les cellules sont en gras (=à modifier)
            if astreinte and ws.cell(row=row, column=2).font.bold:
                jours_gras['garde'].append(row-2)
            if garde and ws.cell(row=row, column=3).font.bold:
                jours_gras['astreinte'].append(row-2)

            # check si les cellules sont soulignées (=à fixer)
            if garde and ws.cell(row=row, column=2).font.underline:
                jours_soulignes['garde'].append(row-2)
            if astreinte and ws.cell(row=row, column=3).font.underline:
                jours_soulignes['astreinte'].append(row-2)
            
            # conversion des noms en indices (on utilise -1 pour les cases vides)
            garde_idx = next((i for i, name in enumerate(mdc) if name == garde), -1) if garde else -1
            astreinte_idx = next((i for i, name in enumerate(mdc) if name == astreinte), -1) if astreinte else -1
            gardes.append(garde_idx)
            astreintes.append(astreinte_idx)

        toutes_cases_remplies = True
        for row in range(2, ws.max_row + 1):
            if row == D+2:
                break

            if (ws.cell(row=row, column=2).value is None or 
                ws.cell(row=row, column=3).value is None):
                toutes_cases_remplies = False
                break
        
        aucune_case_gras = True
        for row in range(2, ws.max_row + 1):
            if row == D+2:
                break

            if (ws.cell(row=row, column=2).font.bold or 
                ws.cell(row=row, column=3).font.bold):
                aucune_case_gras = False
                break
        
        wb.close()

        planning_vide = all(g == -1 for g in gardes) and all(a == -1 for a in astreintes) # si que des -1 -> on a un planning vide
        planning_initial = gardes + astreintes # construction du planning initiale (si vide, on aura ici que des -1)
        skip_optim = not planning_vide and toutes_cases_remplies and aucune_case_gras # condition pour skipper l'optimisation
        
        planning_initiaux.append(planning_initial)
        jours_a_modifier.append(jours_gras)
        jours_fixes.append(jours_soulignes)
        doit_modifier.append(not skip_optim)
        skip_optims.append(skip_optim)
        
        if skip_optim and verbose:
            print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m \033[1m\033[34m{file}\033[0m : Planning complet et sans modification demandée -> pas d'optimisation")

        # tentative de lecture de la feuille "attributs"
        try:
            df_attributs = pd.read_excel(os.path.join(dir, file), sheet_name='attributs')

            mdc_attributs = df_attributs.columns.tolist()[1:]
            if mdc != mdc_attributs: # si les médecins ne correspondent pas entre feuille principale et feuille attributs
                mdc_manquants = set(mdc) - set(mdc_attributs)
                mdc_supplementaires = set(mdc_attributs) - set(mdc)
                
                error_msg = f"\033[1m\033[31m[ERREUR]\033[0m Dans le fichier \033[1m\033[33m{file}\033[0m :"
                if mdc_manquants:
                    error_msg += f"\nMédecins présents dans la feuille principale mais absents de la feuille attributs : \033[1m\033[36m{', '.join(mdc_manquants)}\033[0m"
                if mdc_supplementaires:
                    error_msg += f"\nMédecins présents dans la feuille attributs mais absents de la feuille principale : \033[1m\033[36m{', '.join(mdc_supplementaires)}\033[0m"
                
                print(error_msg)
                return

            attributs_eq = {}

            # parcours de la feuille "attributs"
            for _, row in df_attributs.iterrows():
                nom_attribut = row.iloc[0]
                valeurs = row.iloc[1:].astype(bool).tolist()
                attributs_eq[nom_attribut] = valeurs
                
            attributs_eqs.append(attributs_eq)

            if verbose:
                print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Attributs trouvés pour \033[1m\033[34m{file}\033[0m :")
                for attr, vals in attributs_eq.items():
                    medecins_avec_attr = [m for m, v in zip(mdc, vals) if v]
                    print(f"  - {attr}: {', '.join(medecins_avec_attr)}")
        except ValueError: # si la feuille attributs n'exsite pas pour ce planning, on ajoute quand même {}
            attributs_eqs.append({})

        Ns.append(N)
        Ds.append(D)
        preferences_eqs.append(preferences)
        mdc_eqs.append(mdc)
        global_mdc.update(mdc)

    global_mdc = list(global_mdc)

    # à ce point, on a des indices propres à chaque équipe pour chaque médecin
    # par exemple, si JMR apapartient à 2 équipes, son indice dans la 1ere peut être 6 mais dans la 2e ça peut être 2
    # donc, ici, on construit deux objets eqs_to_global et global_to_eqs pour passer d'un indice exprimé dans une équipe à un indice exprimé dans une autre
    # pour cela, on passe par les indices "global"
    for mdc_eq in mdc_eqs:
        eq_to_global = {i: global_mdc.index(member) for i, member in enumerate(mdc_eq)}
        global_to_eq = {global_mdc.index(member): i for i, member in enumerate(mdc_eq)}

        eqs_to_global.append(eq_to_global)
        global_to_eqs.append(global_to_eq)

    # ici, on calcule pour chaque médecin le nombre d'équipes auquel il appartient, et on stocke ça dans reductions.
    # cela permet, si jamais on n'a pas d'implication donnée, de réduire l'implication de médecins qui sont dans plusieurs équipes
    # cf FAQ pour plus de détails
    nombre_equipes_global = {idx: sum([1 for mdc_eq in mdc_eqs if member in mdc_eq]) for idx, member in enumerate(global_mdc)}

    for mdc_eq, global_to_eq in zip(mdc_eqs, global_to_eqs):
        reductions = [None] * len(mdc_eq)
        for mdc in nombre_equipes_global:
            if mdc in global_to_eq:
                reductions[global_to_eq[mdc]] = nombre_equipes_global[mdc]
        reductions_eqs.append(np.array(reductions))

    # nouvelle lecture des fichiers, pour lire la feuille "implications"
    for file in excel_files:
        df = pd.read_excel(os.path.join(dir, file))
        df = df.drop(df.columns[:3], axis=1)
        df = df.fillna(0)

        mdc = df.columns.to_list()
        preferences = df.to_numpy().T
        N, D = preferences.shape

        implications = {
            'gardes': np.full(N, np.nan),
            'astreintes': np.full(N, np.nan)
        }

        # tentative de lecture de la feuille "implications"
        try:
            df_implications = pd.read_excel(os.path.join(dir, file), sheet_name='implications')

            mdc_implications = df_implications.columns.tolist()[1:]

            # si les médecins ne correspondent pas (idem que "attributs")
            if mdc != mdc_implications:
                mdc_manquants = set(mdc) - set(mdc_implications)
                mdc_supplementaires = set(mdc_implications) - set(mdc)
                
                error_msg = f"\033[1m\033[31m[ERREUR]\033[0m Dans le fichier \033[1m\033[33m{file}\033[0m :"
                if mdc_manquants:
                    error_msg += f"\nMédecins présents dans la feuille principale mais absents de la feuille implications : \033[1m\033[36m{', '.join(mdc_manquants)}\033[0m"
                if mdc_supplementaires:
                    error_msg += f"\nMédecins présents dans la feuille implications mais absents de la feuille principale : \033[1m\033[36m{', '.join(mdc_supplementaires)}\033[0m"
                
                print(error_msg)
                return
            
            # on essaie de lire les valeurs des lignes "gardes" et "astreintes"
            # pour plus de flexibilité, on va localiser ces lignes (plutot que de supposer que gardes est en 1er et astreintes en 2nd)
            indices = {
                'gardes': None,
                'astreintes': None
            }
            for idx, row in df_implications.iterrows():
                if str(row.iloc[0]).lower().strip() == 'gardes':
                    indices['gardes'] = idx
                elif str(row.iloc[0]).lower().strip() == 'astreintes':
                    indices['astreintes'] = idx

            implications['gardes'] = np.full(N, np.nan)
            implications['astreintes'] = np.full(N, np.nan)

            if indices['gardes'] is not None:
                implications['gardes'] = df_implications.iloc[indices['gardes'], 1:].astype(float).values
            if indices['astreintes'] is not None:
                implications['astreintes'] = df_implications.iloc[indices['astreintes'], 1:].astype(float).values

        except ValueError:
            pass

        # gestion des trous dans les implications (incluant le cas où tout est vide) (cf FAQ)
        # pour les astreintes
        masque_trous_astreintes = np.isnan(implications['astreintes'])
        if masque_trous_astreintes.any():
            nb_astreintes_manquantes = D - np.nansum(implications['astreintes'])
            nb_medecins_sans_astreintes = np.sum(masque_trous_astreintes)
            # distribution équitable des astreintes manquantes
            implications['astreintes'][masque_trous_astreintes] = nb_astreintes_manquantes / nb_medecins_sans_astreintes

        # Pour les gardes
        masque_trous_gardes = np.isnan(implications['gardes'])
        # on va calculer un poids pour chaque médecin, avec lequel on va pondéré les gardes qu'on lui donne (en nombre cible)
        if masque_trous_gardes.any():
            nb_gardes_manquantes = D - np.nansum(implications['gardes'])
            # calculer les poids basés sur les préférences positives
            poids = (preferences > 0).sum(axis=1)
            poids_medecins_sans_gardes = poids[masque_trous_gardes]
            if poids_medecins_sans_gardes.sum() > 0:
                poids_medecins_sans_gardes = poids_medecins_sans_gardes / poids_medecins_sans_gardes.sum()
            else:
                # si aucune préférence positive, distribution uniforme
                poids_medecins_sans_gardes = 1/reductions[masque_trous_gardes]
                poids_medecins_sans_gardes = np.ones_like(poids_medecins_sans_gardes) / len(poids_medecins_sans_gardes)
            implications['gardes'][masque_trous_gardes] = nb_gardes_manquantes * poids_medecins_sans_gardes

        # normalisation (il faut que la somme des gardes cibles soit effectivement égale à la somme des gardes à attribuer)
        implications['gardes'] = implications['gardes'] * (D / implications['gardes'].sum())
        implications['astreintes'] = implications['astreintes'] * (D / implications['astreintes'].sum()) # idem pour les astreintes

        implications_eqs.append(implications)

    # des prints
    for i, (planning_initial, jours_gras, doit_modif) in enumerate(zip(planning_initiaux, jours_a_modifier, doit_modifier)):
        if not doit_modif:
            continue
        
        if not jours_gras['garde'] and not jours_gras['astreinte']:
            continue
        
        print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Modifications demandées pour \033[1m\033[34mÉquipe {i+1}\033[0m:")

        if jours_gras['garde']:
            print("  - Gardes à modifier aux jours :", end=" ")
            print(", ".join([f"\033[1m{j+1}\033[0m" for j in jours_gras['garde']]))
            
        if jours_gras['astreinte']:
            print("  - Astreintes à modifier aux jours :", end=" ")
            print(", ".join([f"\033[1m{j+1}\033[0m" for j in jours_gras['astreinte']]))

    if any(bool(jours_gras['garde']) or bool(jours_gras['astreinte']) for jours_gras in jours_a_modifier if jours_gras):
        print()

    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m \033[1m{len(global_mdc)}\033[0m médecins uniques trouvés: ", end="")
    print(", ".join([f"\033[1m\033[36m{mdc}\033[0m" for mdc in global_mdc]))

    for i, (mdc_eq, file) in enumerate(zip(mdc_eqs, excel_files), 1):
        print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m \033[1m\033[34mÉquipe {i}\033[0m (\033[1m\033[33m{file}\033[0m) : ", end="")
        print(", ".join([f"\033[1m\033[36m{mdc}\033[0m" for mdc in mdc_eq]))

    print()
    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Début de l'optimisation.")

    # LANCEMENT DE L'OPTIMISATION : 
    # -on passe toutes les donénes qu'on vient de lire.
    # -on reçoit les plannings et les scores finaux.
    resultat_eqs, score_final_eqs = solve_multi(Ns, Ds, preferences_eqs, reductions_eqs, attributs_eqs, implications_eqs, eqs_to_global, global_to_eqs, planning_initiaux, jours_a_modifier, jours_fixes, skip_optims)

    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Scores finaux par équipe : ", end="")
    print(", ".join([f"\033[1m\033[34mÉquipe {i+1}\033[0m : \033[1m\033[36m{score:.0f}\033[0m" for i, score in enumerate(score_final_eqs)]))

    time.sleep(1)

    print()
    # affichage des modifications par rapport au planning initial (si il y a)
    for i, (resultat_eq, planning_initial, doit_modif) in enumerate(zip(resultat_eqs, planning_initiaux, doit_modifier)):
        if not doit_modif:
            continue
            
        # calcule de la distance (en ne prenant en compte que les cases non vides du planning initial)
        masque = np.array(planning_initial) != -1
        distance = np.sum(np.array(resultat_eq)[masque] != np.array(planning_initial)[masque])
        
        if distance > 0:
            print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m \033[1m\033[34mÉquipe {i+1}\033[0m : \033[1m{distance}\033[0m changements par rapport au planning initial")
        
        if verbose:
            D = Ds[i]
            for jour in range(D):
                # changements de garde
                if planning_initial[jour] != -1 and resultat_eq[jour] != planning_initial[jour]:
                    ancien_mdc = mdc_eqs[i][planning_initial[jour]]
                    nouveau_mdc = mdc_eqs[i][resultat_eq[jour]]
                    print(f"  - Jour {jour+1}, garde : \033[1m\033[36m{ancien_mdc}\033[0m → \033[1m\033[36m{nouveau_mdc}\033[0m")
                
                # changements d'astreinte
                if planning_initial[D + jour] != -1 and resultat_eq[D + jour] != planning_initial[D + jour]:
                    ancien_mdc = mdc_eqs[i][planning_initial[D + jour]]
                    nouveau_mdc = mdc_eqs[i][resultat_eq[D + jour]]
                    print(f"  - Jour {jour+1}, astreinte : \033[1m\033[36m{ancien_mdc}\033[0m → \033[1m\033[36m{nouveau_mdc}\033[0m")

    time.sleep(1)

    # fonction qui check les contraintes dures ainsi que les préférences attribuées
    # et print les jours où des problèmes sont detectés (ils seront aussi affichés dans le Excel, voir plus tard)
    check_coherence(resultat_eqs, Ds, global_mdc, eqs_to_global)

    time.sleep(1)

    # ici on parcours les plannings et on signales les "jours problématiques"
    # (ie les jours où une préférence négative a été attribuée)
    total_jours_problemes = 0
    for i, (resultat_eq, preferences_eq) in enumerate(zip(resultat_eqs, preferences_eqs)):
        jours_problemes_eq = 0
        for jour in range(Ds[i]):
            mdc_garde = resultat_eq[jour]
            preference = preferences_eq[mdc_garde, jour]

            if preference < 0:
                jours_problemes_eq += 1
                total_jours_problemes += 1
                if verbose:
                    if total_jours_problemes == 1:
                        print(f"\033[1m\033[31m[GARDIEN]\033[0m Jours problématiques (médecins affectés avec préférence négative) :")
                        print(f"\033[1m\033[34mÉquipe {i+1}\033[0m :")
                    print(f"  - Jour {jour+1}: Médecin \033[1m\033[36m{mdc_eqs[i][mdc_garde]}\033[0m affecté avec préférence de \033[1m\033[31m{preference}\033[0m")

    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Vérification des jours problématiques...")
    if total_jours_problemes > 0:
        print(f"\033[1m\033[31m[GARDIEN]\033[0m Total de jours problématiques (préférence non respectée): \033[1m\033[31m{total_jours_problemes}\033[0m")
    else:
        print(f"\033[1m\033[32m[GARDIEN]\033[0m Total de jours problématiques (préférence non respectée): \033[1m\033[32m{total_jours_problemes}\033[0m")

    # ici on va parcourir les plannings pour vérifier que les attributs sont représentés chaque jour
    # si ce n'est pas le cas, on signe le jour fautif
    time.sleep(1)
    print()
    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Vérification des attributs des médecins...")

    total_jours_sans_attributs = 0
    for i, (resultat_eq, attributs_eq) in enumerate(zip(resultat_eqs, attributs_eqs)):
        if not attributs_eq:  # Si pas d'attributs pour cette équipe, on passe
            continue

        jours_problemes_eq = 0
        for jour in range(Ds[i]):
            mdc_garde = resultat_eq[jour]
            mdc_astreinte = resultat_eq[Ds[i] + jour]
            
            for nom_attribut, mdc_avec_attribut in attributs_eq.items():
                if not (mdc_avec_attribut[mdc_garde] or mdc_avec_attribut[mdc_astreinte]):
                    jours_problemes_eq += 1
                    total_jours_sans_attributs += 1
                    if verbose:
                        if total_jours_sans_attributs == 1:
                            print(f"\033[1m\033[34mÉquipe {i+1}\033[0m :")
                        print(f"  - Jour {jour+1}: Attribut \033[1m\033[35m{nom_attribut}\033[0m manquant")
                        print(f"    Garde: \033[1m\033[36m{mdc_eqs[i][mdc_garde]}\033[0m")
                        print(f"    Astreinte: \033[1m\033[36m{mdc_eqs[i][mdc_astreinte]}\033[0m")

    if total_jours_sans_attributs > 0:
        print(f"\033[1m\033[31m[GARDIEN]\033[0m Total de jours avec attributs manquants: \033[1m\033[31m{total_jours_sans_attributs}\033[0m")
    else:
        print(f"\033[1m\033[32m[GARDIEN]\033[0m Tous les jours ont les attributs requis.")

    time.sleep(1)
    print()
    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Enregistrement des plannings en cours...")

    # début de la phase d'exportation : un fichier par équipe
    blue_fill = PatternFill(start_color="6c9beb", end_color="6c9beb", fill_type="solid")
    red_fill = PatternFill(start_color="DB7C6C", end_color="DB7C6C", fill_type="solid")
    green_fill = PatternFill(start_color="99c57a", end_color="99c57a", fill_type="solid")
    gray_fill = PatternFill(start_color="9a9796", end_color="9a9796", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
    collision_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    jour_off_fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
    attribut_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
    modification_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    nonassigne_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

    # check des collisions, problèmes de jour OFF, problèmes d'attributs manquants
    schedule = np.zeros((len(global_mdc), max(Ds)), dtype=int) # planning global en fonction des médecins uniques
    # (0=rien, 1= garde, 2=astreinte; pour chaque médecin "unique" ie la réunion sans répétition des médecins de chaque planning)
    collisions = {i: set() for i in range(len(excel_files))} # jours avec collision par équipe
    jours_off = {i: set() for i in range(len(excel_files))}  # jours avec problème de jour off
    jours_attributs = {i: set() for i in range(len(excel_files))} # jours avec attribut manquant

    for i, (resultat_eq, eq_to_global) in enumerate(zip(resultat_eqs, eqs_to_global)):
        planning_gardes = resultat_eq[:Ds[i]]
        planning_astreintes = resultat_eq[Ds[i]:]

        # collisions ?
        for day, mdc_local in enumerate(planning_gardes):
            if mdc_local == -1: # case non remplie par l'optimisation: on passe
                continue
            mdc_global = eq_to_global[mdc_local]
            if schedule[mdc_global, day] != 0:
                collisions[i].add(day)
            schedule[mdc_global, day] = 1

        # collisions ?
        for day, mdc_local in enumerate(planning_astreintes):
            if mdc_local == -1: # case non remplie par l'optimisation: on passe
                continue
            mdc_global = eq_to_global[mdc_local]
            if schedule[mdc_global, day] != 0:
                collisions[i].add(day)
            schedule[mdc_global, day] = 2

        # jours off ok?
        for day in range(Ds[i] - 1):
            if planning_gardes[day] == planning_gardes[day + 1] or \
            planning_gardes[day] == planning_astreintes[day + 1]:
                jours_off[i].add(day + 1)

        # attributs ok ?
        if attributs_eqs[i]:
            for jour in range(Ds[i]):
                mdc_garde = planning_gardes[jour]
                mdc_astreinte = planning_astreintes[jour]
                for nom_attribut, mdc_avec_attribut in attributs_eqs[i].items():
                    if not (mdc_avec_attribut[mdc_garde] or mdc_avec_attribut[mdc_astreinte]):
                        jours_attributs[i].add(jour)

    fichiers_crees = []
    for i, file in enumerate(excel_files):
        wb = load_workbook(os.path.join(dir, file))

        # création de la feuille où l'on place le résumé

        # suppression 
        if "résumé" in wb.sheetnames:
            del wb["résumé"]
        ws_resume = wb.create_sheet("résumé")

        headers = ["Médecin", "Gardes effectuées", "Gardes ciblées", "Astreintes effectuées", 
               "Astreintes ciblées", "Écart moyen entre gardes", "Gardes sur préf. négatives", 
               "Gardes sur préf. nulles"]
        for col, header in enumerate(headers, 1):
            cell = ws_resume.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            # Colorer les entêtes des colonnes de préférences
            if header == "Gardes sur préf. négatives":
                cell.fill = yellow_fill
            elif header == "Gardes sur préf. nulles":
                cell.fill = yellow_fill

        # on va récupérer les données du planning pour faire un "résumé" : combien de gardes, etc...
        # (ce résumé concerne CE planning seulement, pas le planning global sur plusieurs équipes)
        resultat_eq = resultat_eqs[i]
        planning_gardes = resultat_eq[:Ds[i]]
        planning_astreintes = resultat_eq[Ds[i]:]
        
        # on parcourt les médecins de l'équipe
        for j, mdc_name in enumerate(mdc_eqs[i]):
            # on compte les gardes et astreintes
            nb_gardes = sum(1 for x in planning_gardes if x == j)
            nb_astreintes = sum(1 for x in planning_astreintes if x == j)
            
            # on récupère les valeurs cibles
            target_gardes = implications_eqs[i]['gardes'][j]
            target_astreintes = implications_eqs[i]['astreintes'][j]
            
            # écart moyen entre les gardes
            jours_garde = np.where(planning_gardes == j)[0]
            if len(jours_garde) > 1:
                ecart_moyen = np.mean(np.diff(jours_garde))
            else:
                ecart_moyen = "-"
                
            # nombre de gardes sur préférences négatives ou nulles
            gardes_pref_neg = sum(1 for day in range(Ds[i]) if planning_gardes[day] == j 
                                and preferences_eqs[i][j][day] < 0)
            gardes_pref_null = sum(1 for day in range(Ds[i]) if planning_gardes[day] == j 
                                and preferences_eqs[i][j][day] == 0)
            
            # remplissage de la ligne
            row = j + 2
            ws_resume.cell(row=row, column=1, value=mdc_name)
            ws_resume.cell(row=row, column=2, value=nb_gardes)
            # Pour les nombres avec décimales, on garde les nombres et on formate la cellule
            cell_target_gardes = ws_resume.cell(row=row, column=3, value=target_gardes)
            cell_target_gardes.number_format = '#,##0.0'  # Format français avec 1 décimale
            ws_resume.cell(row=row, column=4, value=nb_astreintes)
            cell_target_astreintes = ws_resume.cell(row=row, column=5, value=target_astreintes)
            cell_target_astreintes.number_format = '#,##0.0'
            if isinstance(ecart_moyen, float):
                cell_ecart = ws_resume.cell(row=row, column=6, value=ecart_moyen)
                cell_ecart.number_format = '#,##0.0'
            else:
                ws_resume.cell(row=row, column=6, value=ecart_moyen)
            ws_resume.cell(row=row, column=7, value=gardes_pref_neg)
            ws_resume.cell(row=row, column=8, value=gardes_pref_null)
        
        for col in range(1, len(headers) + 1):
            ws_resume.column_dimensions[get_column_letter(col)].width = 25 # largeur colonne (esthétique)

        # création de la feuille où l'on place les légendes
        # c'est essentiellement du formatage/esthétisme
        # suppression 
        if "légende" in wb.sheetnames:
            del wb["légende"]
        ws_legende = wb.create_sheet("légende")

        title_font = Font(bold=True, size=12)
        subtitle_font = Font(bold=True)

        # titre
        ws_legende['A1'] = "LÉGENDE DES COULEURS"
        ws_legende['A1'].font = title_font
        
        # légende des problèmes "graves"
        ws_legende['A3'] = "Problèmes graves :"
        ws_legende['A3'].font = subtitle_font
        
        ws_legende['B4'] = "Collision (médecin assigné à plusieurs équipes)"
        ws_legende['A4'].fill = collision_fill
        
        ws_legende['B5'] = "Jour OFF après garde non respecté"
        ws_legende['A5'].fill = jour_off_fill
        
        ws_legende['B6'] = "Attribut manquant ce jour-là"
        ws_legende['A6'].fill = attribut_fill

        ws_legende['B7'] = "Garde ou astreinte non attribuée"
        ws_legende['A7'].fill = nonassigne_fill

        # légende des préférences
        ws_legende['A9'] = "Préférences :"
        ws_legende['A9'].font = subtitle_font
        
        ws_legende['B10'] = "Préférence positive ou nulle (assignée à une garde)"
        ws_legende['A10'].fill = blue_fill
        
        ws_legende['B11'] = "Préférence négative (assignée à une garde)"
        ws_legende['A11'].fill = yellow_fill
        
        ws_legende['B12'] = "Préférence positive"
        ws_legende['A12'].fill = green_fill
        
        ws_legende['B13'] = "Préférence négative"
        ws_legende['A13'].fill = red_fill
        
        ws_legende['B14'] = "Préférence nulle"
        ws_legende['A14'].fill = gray_fill

        # légende des modifications
        ws_legende['A16'] = "Modifications :"
        ws_legende['A16'].font = subtitle_font
        
        ws_legende['B17'] = "Garde ou astreinte modifiée par rapport au planning initial"
        ws_legende['A17'].fill = modification_fill

        ws_legende.column_dimensions['B'].width = 50
        ws_legende.column_dimensions['A'].width = 4

        for k in range(1, 18):
            ws_legende.row_dimensions[k].height = 20

        # modification de la feuille principale (remplissage colonne garde+astreinte, colorations)
        ws = wb.worksheets[0]

        for day in range(Ds[i]):
            mdc_garde = resultat_eqs[i][day]
            mdc_astreinte = resultat_eqs[i][Ds[i]:][day]

            # on récupere les cellules correspondante au jour, garde, astreinte de la ligne
            garde_cell = ws.cell(row=day+2, column=2)
            astreinte_cell = ws.cell(row=day+2, column=3)
            jour_cell = ws.cell(row=day+2, column=1)

            # coloration des gardes/astreintes si problèmes détectés
            if day in collisions[i]:
                jour_cell.fill = collision_fill
            if day in jours_off[i]:
                jour_cell.fill = jour_off_fill
            if day in jours_attributs[i]:
                jour_cell.fill = attribut_fill
            if (mdc_garde==-1) or (mdc_astreinte==-1): # case vide: on le signale
                jour_cell.fill = nonassigne_fill

            # coloration des changements si une modification par rapport au planning initial
            if planning_initiaux[i] is not None:
                if planning_initiaux[i][day] != -1 and mdc_garde != planning_initiaux[i][day]:
                    garde_cell.fill = modification_fill
                if planning_initiaux[i][Ds[i] + day] != -1 and mdc_astreinte != planning_initiaux[i][Ds[i] + day]:
                    astreinte_cell.fill = modification_fill
            
            # remplir et colorer les préférences
            for j in range(Ns[i]):
                pref_cell = ws.cell(row=day+2, column=4+j)
                preference = pref_cell.value
                
                if preference is None:
                    pref_cell.fill = gray_fill
                elif preference < 0:
                    if j == mdc_garde:
                        pref_cell.fill = yellow_fill # pref négative assignée
                    else:
                        pref_cell.fill = red_fill # pref négative non assignée
                elif preference >= 0:
                    if j == mdc_garde:
                        pref_cell.fill = blue_fill # pref positive assignée
                    else:
                        pref_cell.fill = green_fill # pref positive non assignée

            # update colonne "garde" et "astreinte" pour ce jour
            garde_cell.value = mdc_eqs[i][mdc_garde] if mdc_garde != -1 else ""
            astreinte_cell.value = mdc_eqs[i][mdc_astreinte] if mdc_astreinte != -1 else ""

            # on enlève le gras de garde et astreinte
            # (gras=valeur à modifier, or, ça y est on a modifié)
            garde_cell.font = Font(bold=False)
            astreinte_cell.font = Font(bold=False)
            ws.cell(row=day+2, column=1).font = Font(bold=False)

            # par contre, on garde le soulignage pour les valeurs fixés
            if day in jours_fixes[i]['garde']:
                garde_cell.font = Font(underline='single')
            if day in jours_fixes[i]['astreinte']:
                astreinte_cell.font = Font(underline='single')

        if mode == 'R':
            output_filename = os.path.join(dir, file)
        else:
            output_filename = os.path.join(dir, f"{os.path.splitext(file)[0]}_resultat.xlsx")
        wb.save(output_filename)
        wb.close()
        fichiers_crees.append(os.path.basename(output_filename))

    print(f"\033[1m\033[32m[GARDIEN]\033[0m Enregistrement réussi: \033[1m\033[33m{', '.join([f'{fichier}' for fichier in fichiers_crees])}\033[0m")

def check_coherence(resultat_eqs, Ds, global_mdc, eqs_to_global):
    """
    Vérification des collisions et du respect des jours OFF après les gardes
    """
   
    print()
    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Vérification des collisions et des jours OFF après les gardes...")

    schedule = np.zeros((len(global_mdc), max(Ds)), dtype=int)
    detecte = False

    for i, (resultat_eq, eq_to_global) in enumerate(zip(resultat_eqs, eqs_to_global)):
        planning_gardes = resultat_eq[:Ds[i]]  # première moitié pour les gardes
        planning_astreintes = resultat_eq[Ds[i]:] # seconde moitié pour les astreintes

        # on remplit les gardes
        for day, mdc_local in enumerate(planning_gardes):
            if mdc_local == -1: # case non remplie par l'optimisation: on passe (problématique)
                continue
            mdc_global = eq_to_global[mdc_local] # conversion en ID global du mdc

            if schedule[mdc_global, day] != 0:
                print(f"\033[1m\033[31m[ERREUR]\033[0m Collision pour le médecin \033[1m\033[36m{global_mdc[mdc_global]}\033[0m au jour {day+1}")
                detecte = True
            schedule[mdc_global, day] = 1 # 1=garde

        # on remplit les astreintes
        for day, mdc_local in enumerate(planning_astreintes):
            if mdc_local == -1: # case non remplie par l'optimisation: on passe (problématique)
                continue
            mdc_global = eq_to_global[mdc_local] # conversion en ID global du mdc

            if schedule[mdc_global, day] != 0:
                print(f"\033[1m\033[31m[ERREUR]\033[0m Collision pour le médecin \033[1m\033[36m{global_mdc[mdc_global]}\033[0m au jour {day+1}")
                detecte = True
            schedule[mdc_global, day] = 2 # 2=astreinte

    # vérification jour OFF
    for mdc_global in range(len(global_mdc)):
        for day in range(max(Ds) - 1):
            if schedule[mdc_global, day] == 1 and schedule[mdc_global, day + 1] != 0:
                print(f"\033[1m\033[31m[ERREUR]\033[0m Le médecin \033[1m\033[36m{global_mdc[mdc_global]}\033[0m travaille le jour {day+2} après une garde au jour {day+1}")
                detecte = True

    if not detecte:
        print("\033[1m\033[32m[GARDIEN]\033[0m Aucun problème de collision ou de jour OFF détecté.")
    print()

    time.sleep(1)
    print(f"\033[1m{couleur_gardien}[GARDIEN]\033[0m Vérification des gardes/astreintes non attribuées...")

    total_non_attribues = 0
    for i, resultat_eq in enumerate(resultat_eqs):
        jours_non_attribues_eq = {'garde': [], 'astreinte': []}
        
        # pour les gardes
        for jour, mdc in enumerate(resultat_eq[:Ds[i]]):
            if mdc == -1: # -1 = case vide (non remplie par l'optimisation)
                total_non_attribues += 1
                jours_non_attribues_eq['garde'].append(jour + 1)
        
        # pour les astreintes        
        for jour, mdc in enumerate(resultat_eq[Ds[i]:]):
            if mdc == -1:
                total_non_attribues += 1
                jours_non_attribues_eq['astreinte'].append(jour + 1)
                
        if jours_non_attribues_eq['garde'] or jours_non_attribues_eq['astreinte']:
            print(f"\033[1m\033[34mÉquipe {i+1}\033[0m :")
            if jours_non_attribues_eq['garde']:
                print(f"  - Gardes non attribuées aux jours : {', '.join(map(str, jours_non_attribues_eq['garde']))}")
            if jours_non_attribues_eq['astreinte']:
                print(f"  - Astreintes non attribuées aux jours : {', '.join(map(str, jours_non_attribues_eq['astreinte']))}")

    if total_non_attribues > 0:
        print(f"\033[1m\033[31m[GARDIEN]\033[0m Total de gardes/astreintes non attribuées: \033[1m\033[31m{total_non_attribues}\033[0m")
    else:
        print(f"\033[1m\033[32m[GARDIEN]\033[0m Toutes les gardes et astreintes ont été attribuées.")
    print()

if __name__ == "__main__":
    print_ascii()
    time.sleep(1)
    main()
